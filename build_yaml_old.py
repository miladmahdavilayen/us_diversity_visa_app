import openpyxl
import yaml


def read_yaml(full_name):
    with open(f'data/yaml_forms/{full_name}.yaml', 'r') as file:
        yaml_data = yaml.safe_load(file)
    return yaml_data


def build_yaml(person_name):
    # Load the Excel file
    workbook = openpyxl.load_workbook(f'data/forms/{person_name}.xlsx')  # Replace with the actual file path
    sheet = workbook.active

    # Create a dictionary to store the parameter data
    parameters = {
        "first_name": sheet.cell(row=2, column=2).value,
        "last_name": sheet.cell(row=3, column=2).value,
        "bday_day": sheet.cell(row=4, column=2).value,
        "bday_month": sheet.cell(row=5, column=2).value,
        "bday_year": sheet.cell(row=6, column=2).value,
        "country_born": sheet.cell(row=7, column=2).value,  # Updated key
        "city_born": sheet.cell(row=8, column=2).value,  # Updated key
        "address_1": sheet.cell(row=9, column=2).value,  # Updated key
        "address_city": sheet.cell(row=10, column=2).value,  # Updated key
        "province": sheet.cell(row=11, column=2).value,
        "zip_code": sheet.cell(row=12, column=2).value,
        "country_mail_live": sheet.cell(row=13, column=2).value,  # Updated key
        "phone": sheet.cell(row=14, column=2).value,
        "email": sheet.cell(row=15, column=2).value,
        "gender": sheet.cell(row=16, column=2).value,
        "education": sheet.cell(row=17, column=2).value,
        "marital_status": sheet.cell(row=18, column=2).value,
        "spouce_full_name": sheet.cell(row=19, column=2).value,
        "num_kids": sheet.cell(row=20, column=2).value,  # Updated key
        "no_zip_code": False
    }

    # Create the YAML parameter file
    full_name = parameters['first_name'] + parameters['last_name']
    with open(f'data/yaml_forms/{full_name}.yaml', 'w') as file:
        yaml.dump(parameters, file, default_flow_style=False)

    return read_yaml(full_name)